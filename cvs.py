import datetime
import sys
import pandas as pd
import numpy as np
import os

from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill

roasFile = "./cvs/roas.csv"
retentionFile = "./cvs/retention.csv"
revenueFile = "./cvs/revenue.csv"
roasSortFile = "./cvs/roas_sort.csv"
retentionSortFile = "./cvs/retention_sort.csv"
revenueSortFile = "./cvs/revenue_sort.csv"
# clearBeforeFile = "./cvs/clear_before.csv"
# clearAfterFile = "./cvs/clear_after.csv"
resultCSVFile = "./cvs/result.csv"
resultExcelFile = "./cvs/result.xlsx"

retentionDay1User = 'sessions - Unique users - day 1 - partial'
roasRetentionDay1 = "retention day1 "
roasRetentionDay1Percent = "retention day1 percent"


def sort(df, f):
    df.sort_values(by=['Media Source', 'Campaign', 'Campaign Id', 'Cohort Day'], inplace=True, ascending=True)
    # df.to_csv(f, index=False)
    return df


def format_float(df):
    value = round(df * 100, 2)
    return value


def calculate_row(df_row):
    return str((format_float(df_row[retentionDay1User] / df_row['Users']))) + "%"


# 计算day0留存率
def retention_day0_percent(df, f):
    # 计算 sessions - Unique users - day 1 列和 Users 列的百分比

    drop_column(df, roasRetentionDay1Percent)
    df[roasRetentionDay1Percent] = df.apply(lambda df_row: calculate_row(df_row), axis=1)


def clear_table(df, f, v, days, end_date):
    # 遍历每一行
    for index, row in df.iterrows():
        # 获取该行的Cohort Day
        n_day_str = row['Cohort Day']
        n_day = datetime.strptime(n_day_str, '%Y-%m-%d')
        # print("start "+n_day_str)

        # 循环30次
        for i in range(days):
            # 获取临时字符串变量temp
            temp = f(i)

            # 获取日期
            column_day = n_day + timedelta(days=i)
            # print("column_day_str " + column_day_str + " curr_day " + curr_day_str)

            # 如果 column_day 大于 curr_day，且该行的 temp 列存在，则清空该行的 temp 列
            if column_day > end_date and temp in df.columns:
                df.at[index, temp] = v
                # print("clear column_day_str " + column_day_str)


# 删除df的f列
def drop_column(df, f):
    if f in df.columns:
        df.drop(f, axis=1, inplace=True)


# 获取roas表格day n列名字
def get_roas_column(day_str):
    return f"roas - Rate - day {day_str} - partial"


# 获取revenue表格day n名字
def get_revenue_column(day_str):
    return f"revenue - sum - day {day_str} - partial"


def sort_all_table(roas_df, retention_df, revenue_df):
    sort(roas_df, roasSortFile)
    sort(retention_df, retentionSortFile)
    sort(revenue_df, revenueSortFile)


# retention表中day0用户数和留存率添加到roas
def merge_retention_to_roas(roas_df, retention_df):
    # 计算day0留存率
    retention_day0_percent(retention_df, retentionFile)
    # 获取day0用户数
    retention_day0_user_column = retention_df[retentionDay1User]

    # 获取day0用户留存率
    retention_day0_user_percent_column = retention_df[roasRetentionDay1Percent]

    drop_column(roas_df, roasRetentionDay1)
    # 将retention_df表的索引重置
    roas_df.insert(loc=7, column=roasRetentionDay1, value=retention_day0_user_column)

    drop_column(roas_df, roasRetentionDay1Percent)
    roas_df.insert(loc=8, column=roasRetentionDay1Percent, value=retention_day0_user_percent_column)


# roas添加day n收入
def merge_revenue_to_roas(roas_df, revenue_df):
    # 提取包含关键字的列
    revenue_cols = [col for col in revenue_df.columns if 'revenue - sum - day' in col.lower()]

    # 将提取的列插入到 roas_df 中
    for col in revenue_cols:
        roas_df[col] = revenue_df[col]


def clear_roas(roas_df, days, end_date):
    # roas_df.to_csv(clearBeforeFile, index=False)
    clear_table(roas_df, get_roas_column, np.NAN, days, end_date)
    # clear_table(roas_df, get_revenue_column, 0)
    # roas_df.to_csv(clearAfterFile, index=False)


# 定义排序函数
def sort_table(table):
    # 获取Cohort Day列的最大值和最小值
    max_date = table["Cohort Day"].max()
    min_date = table["Cohort Day"].min()
    # 返回一个元组，用于排序
    return (max_date, min_date)


def split_table(df, days):

    # 分割表格，按Media Source和campaign分组
    groups = df.groupby(["Media Source", "Campaign", "Campaign Id"])

    # 存储分割结果的列表
    split_table_list = []

    # 遍历每个分组
    for group_name, group_data in groups:
        # 将分组结果存入split_table_list
        split_table_list.append(group_data)

    # 对split_table_list排序
    sorted_table_list = sorted(split_table_list, key=sort_table, reverse=True)

    # 在每个表格最后一行插入空行
    for table in split_table_list:
        table.loc[len(table)] = [None] * len(table.columns)

    # # 遍历排序后的表格，并保存到本地
    # for i, table in enumerate(sorted_table_list):
    #     # 将表格保存到本地
    #     filename = f"{i}.csv"
    #     table.to_csv(filename, index=False)

    merged_table = pd.concat(sorted_table_list)
    merged_table = merged_table.reset_index(drop=True)

    # 获取需要重命名的列列表
    old_cols = [get_roas_column(i) for i in range(days)]

    # 构造新列名列表
    new_cols = [f'roas_day{i}' for i in range(days)]

    # 使用字典将旧列名和新列名对应起来，实现批量重命名
    col_rename_dict = dict(zip(old_cols, new_cols))
    merged_table = merged_table.rename(columns=col_rename_dict)

    # 输出到文件
    merged_table.to_csv(resultCSVFile, index=False)

    # 将DataFrame保存为excel文件
    # merged_table.to_excel(resultExcelFile, index=False)
    format_excel(merged_table)


def format_excel(df):
    df.to_excel(resultExcelFile, index=False)

    # 打开Excel文件
    workbook = openpyxl.load_workbook(resultExcelFile)
    worksheet = workbook.active

    retention_day1_percent_key = 'retention day1 percent'
    roas_day0_key = 'roas_day0'
    roas_day7_key = 'roas_day7'
    roas_day14_key = 'roas_day14'
    roas_day21_key = 'roas_day21'
    roas_day30_key = 'roas_day30'
    # 根据列名找到对应的列索引
    if retention_day1_percent_key in df.columns:
        retention_day1_idx = df.columns.get_loc(retention_day1_percent_key) + 1
    if roas_day0_key in df.columns:
        roas_day0_idx = df.columns.get_loc(roas_day0_key) + 1
    if roas_day7_key in df.columns:
        roas_day7_idx = df.columns.get_loc(roas_day7_key) + 1
    if roas_day14_key in df.columns:
        roas_day14_idx = df.columns.get_loc(roas_day14_key) + 1
    if roas_day21_key in df.columns:
        roas_day21_idx = df.columns.get_loc(roas_day21_key) + 1
    if roas_day30_key in df.columns:
        roas_day30_idx = df.columns.get_loc(roas_day30_key) + 1

    # 配置格式
    red_font = Font(color='FFFF99')
    light_pink = Font(color='FFD9EC')
    light_red_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    light_pink_fill = PatternFill(start_color='FFD9EC', end_color='FFD9EC', fill_type='solid')

    # 遍历每一行，对指定列应用格式
    for row in worksheet.iter_rows(min_row=2, min_col=1):
        retention_day1_cell = row[retention_day1_idx - 1]
        roas_day0_cell = row[roas_day0_idx - 1]
        roas_day7_cell = row[roas_day7_idx - 1]
        roas_day14_cell = row[roas_day21_idx - 1]
        roas_day21_cell = row[roas_day14_idx - 1]
        roas_day30_cell = row[roas_day30_idx - 1]

        retention_day1_cell.fill = light_red_fill
        roas_day0_cell.fill = light_pink_fill
        roas_day7_cell.fill = light_pink_fill
        roas_day14_cell.fill = light_pink_fill
        roas_day21_cell.fill = light_pink_fill
        roas_day30_cell.fill = light_pink_fill

        # retention_day1_cell.font = red_font
        # roas_day0_cell.font = light_pink
        # roas_day7_cell.font = light_pink
        # roas_day14_cell.font = light_pink
        # roas_day21_cell.font = light_pink
        # roas_day30_cell.font = light_pink

    # 保存Excel文件
    workbook.save(resultExcelFile)



def main():
    days = 31
    end_date = datetime.now() - timedelta(days=2)
    if len(sys.argv) > 1:
        days = sys.argv[1]
        print("days:", days)
    if len(sys.argv) > 2:
        end_date_str = sys.argv[2]
        end_date = datetime.now().strftime('%Y-%m-%d')
        print("end_date_str:", end_date_str)
    print("end_date:", end_date)
    # 读取 roas.csv 文件
    roas_df = pd.read_csv(roasFile, dtype={'Campaign Id': str})

    # 读取 retention.csv 文件
    retention_df = pd.read_csv(retentionFile, dtype={'Campaign Id': str})

    # 读取 revenue.csv 文件
    revenue_df = pd.read_csv(revenueFile, dtype={'Campaign Id': str})
    sort_all_table(roas_df, retention_df, revenue_df)
    # 重置索引
    roas_df = roas_df.reset_index(drop=True)
    retention_df = retention_df.reset_index(drop=True)
    revenue_df = revenue_df.reset_index(drop=True)
    merge_retention_to_roas(roas_df, retention_df)
    merge_revenue_to_roas(roas_df, revenue_df)
    clear_roas(roas_df, days, end_date)
    split_table(roas_df, days)


# 函数入口
main()















